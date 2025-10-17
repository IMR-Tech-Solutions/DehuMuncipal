from django.shortcuts import render
from rest_framework import status
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from .models import Survey
from .serializers import SurveySerializer
from django.http import HttpResponse
from django.db.models import Q
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import io
from django.db.models.functions import TruncMonth
from django.db.models import Count
from datetime import datetime, timedelta
import logging
from accounts.premissions import HasModuleAccess
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
import pandas as pd
from django.db import transaction

# Set up logging
logger = logging.getLogger(__name__)


# ==============================
# Survey Create View
# ==============================
@method_decorator(csrf_exempt, name="dispatch")
class SurveyCreateView(APIView):
    permission_classes = [IsAuthenticated, HasModuleAccess]
    required_permission = "create-survey"
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request, *args, **kwargs):
        data = request.data.copy()

        ward_no = data.get("ward_no")
        property_no = data.get("property_no")

        # Check duplicate
        if Survey.objects.filter(ward_no=ward_no, property_no=property_no).exists():
            return Response(
                {
                    "success": False,
                    "message": f"Survey with ward_no {ward_no} and property_no {property_no} already exists.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Create survey
        serializer = SurveySerializer(data=data, context={"request": request})
        if serializer.is_valid():
            survey = serializer.save()
            logger.info(f"Survey created successfully: {survey.id}")

            return Response(
                {
                    "success": True,
                    "message": "Survey created successfully!",
                    "survey_id": survey.id,
                    "data": SurveySerializer(survey).data,
                },
                status=status.HTTP_201_CREATED,
            )
        else:
            logger.error(f"Survey validation errors: {serializer.errors}")
            return Response(
                {
                    "success": False,
                    "message": "Failed to create survey",
                    "errors": serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )


# ==============================
# Survey Detail View
# ==============================
class SurveyDetailView(APIView):
    permission_classes = [IsAuthenticated, HasModuleAccess]
    method_permissions = {
        "GET": "access-survey",
        "PUT": "edit-survey",
        "PATCH": "edit-survey",
        "DELETE": "delete-survey",
    }

    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_permissions(self):
        self.required_permission = self.method_permissions.get(
            self.request.method, "access-survey"
        )
        return super().get_permissions()

    def get_object(self, pk):
        try:
            return Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return None

    def get(self, request, pk, format=None):
        survey = self.get_object(pk)
        if survey is None:
            return Response(
                {"message": "Survey not found"}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = SurveySerializer(survey)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk, format=None):
        return self.update_survey(request, pk, partial=False)

    def patch(self, request, pk, format=None):
        return self.update_survey(request, pk, partial=True)

    def update_survey(self, request, pk, partial=False):
        survey = self.get_object(pk)
        if survey is None:
            return Response(
                {"success": False, "message": "Survey not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        data = request.data.copy()

        serializer = SurveySerializer(survey, data=data, partial=partial)
        if not serializer.is_valid():
            return Response(
                {
                    "success": False,
                    "message": "Validation failed",
                    "errors": serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        survey = serializer.save()

        return Response(
            {
                "success": True,
                "message": "Survey updated successfully!",
                "data": SurveySerializer(survey).data,
            },
            status=status.HTTP_200_OK,
        )

    def delete(self, request, pk, format=None):
        survey = self.get_object(pk)
        if survey is None:
            return Response(
                {"success": False, "message": "Survey not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        survey.delete()
        return Response(
            {"success": True, "message": "Survey deleted successfully"},
            status=status.HTTP_204_NO_CONTENT,
        )


# ==============================
# Survey List View
# ==============================
class SurveyListView(APIView):
    permission_classes = [IsAuthenticated, HasModuleAccess]
    required_permission = "access-survey"

    def get(self, request, format=None):
        try:
            surveys = Survey.objects.all()
            paginator = PageNumberPagination()
            paginator.page_size = 10
            result_page = paginator.paginate_queryset(surveys, request)
            serializer = SurveySerializer(result_page, many=True)

            return paginator.get_paginated_response(serializer.data)

        except Exception as e:
            return Response(
                {"success": False, "message": f"Error retrieving surveys: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ==============================
# Excel Export
# ==============================
@method_decorator(csrf_exempt, name="dispatch")
class SurveyExcelExportView(APIView):
    """Export Survey data to Excel"""

    permission_classes = [IsAuthenticated, HasModuleAccess]
    required_permission = "export-survey"

    def post(self, request):
        try:
            # Get parameters from request
            ward_no = request.data.get("ward_no")
            property_no_start = request.data.get("property_no_start")
            property_no_end = request.data.get("property_no_end")

            # Validation
            if not ward_no:
                return Response(
                    {"success": False, "message": "Ward number is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if not property_no_start or not property_no_end:
                return Response(
                    {"success": False, "message": "Property number range is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:
                property_start = int(property_no_start)
                property_end = int(property_no_end)
                ward_number = int(ward_no)
            except ValueError:
                return Response(
                    {"success": False, "message": "Please enter valid numbers"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if property_start > property_end:
                return Response(
                    {
                        "success": False,
                        "message": "Start property number should be less than end property number",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Filter surveys
            surveys = Survey.objects.filter(
                ward_no=ward_number,
                property_no__gte=property_start,
                property_no__lte=property_end,
            ).order_by("property_no")

            if not surveys.exists():
                return Response(
                    {
                        "success": False,
                        "message": f"No data found for Ward {ward_number}, Property No. {property_start}-{property_end}",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Generate Excel file
            excel_file = self.generate_excel(surveys, ward_number, property_start, property_end)

            # Create response
            response = HttpResponse(
                excel_file.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            filename = f'Survey_Ward_{ward_number}_Property_{property_start}_to_{property_end}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response(
                {"success": False, "message": f"Export error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def generate_excel(self, surveys, ward_no, property_start, property_end):
        """Generate Excel file with Survey data"""

        wb = Workbook()
        ws = wb.active
        ws.title = "Survey Data"

        # Headers
        headers = [
            "id",
            "ward_no",
            "property_no",
            "old_ward_no",
            "old_property_no",
            "property_description",
            "address",
            "address_marathi",
            "water_connection_available",
            "number_of_water_connections",
            "connection_size",
            "remarks",
            "remarks_marathi",
            "created_at",
            "updated_at",
        ]

        ws.append(headers)

        # Helper function
        def make_naive(dt):
            if isinstance(dt, datetime):
                return dt.replace(tzinfo=None)
            return dt

        # Write data rows
        for survey in surveys:
            row = []
            for fld in headers:
                val = getattr(survey, fld, "")
                row.append(make_naive(val))
            ws.append(row)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save Excel
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file


# ==============================
# Excel Import
# ==============================
@method_decorator(csrf_exempt, name="dispatch")
class SurveyExcelImportView(APIView):
    """Import Survey data from Excel file"""

    permission_classes = [IsAuthenticated, HasModuleAccess]
    required_permission = "import-survey"
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        try:
            # Excel file upload check
            if "excel_file" not in request.FILES:
                return Response(
                    {"success": False, "message": "Excel file is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            excel_file = request.FILES["excel_file"]

            # Extension check
            if not excel_file.name.endswith((".xlsx", ".xls")):
                return Response(
                    {"success": False, "message": "Only Excel files (.xlsx, .xls) are allowed"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Read Excel
            try:
                df = pd.read_excel(excel_file, engine="openpyxl")
                logger.info(f"Excel file read successfully. Columns: {df.columns.tolist()}")
                logger.info(f"Total rows: {len(df)}")
            except Exception as e:
                return Response(
                    {"success": False, "message": f"Error reading Excel file: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Process and import
            result = self.process_excel_data(df, request)

            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Excel import error: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {"success": False, "message": f"Import error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def process_excel_data(self, df, request):
        """Process Excel data and import to database"""
        success_count = 0
        error_count = 0
        errors = []

        # Required columns
        required_columns = ["ward_no", "property_no"]

        # Check required columns
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return {
                "success": False,
                "message": f"Required columns missing: {', '.join(missing_columns)}",
                "missing_columns": missing_columns,
            }

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    ward_no = row.get("ward_no")
                    property_no = row.get("property_no")

                    # Duplicate check
                    if Survey.objects.filter(ward_no=ward_no, property_no=property_no).exists():
                        error_count += 1
                        errors.append(
                            f"Row {index+2}: Ward {ward_no}, Property {property_no} - Survey already exists"
                        )
                        continue

                    # Prepare survey data
                    survey_data = {
                        "ward_no": int(ward_no) if pd.notna(ward_no) else None,
                        "property_no": str(property_no) if pd.notna(property_no) else None,
                        "old_ward_no": str(row.get("old_ward_no", "")) if pd.notna(row.get("old_ward_no")) else "",
                        "old_property_no": str(row.get("old_property_no", "")) if pd.notna(row.get("old_property_no")) else "",
                        "property_description": str(row.get("property_description", "")) if pd.notna(row.get("property_description")) else "",
                        "address": str(row.get("address", "")) if pd.notna(row.get("address")) else "",
                        "address_marathi": str(row.get("address_marathi", "")) if pd.notna(row.get("address_marathi")) else "",
                        "water_connection_available": str(row.get("water_connection_available", "No")),
                        "number_of_water_connections": int(row.get("number_of_water_connections", 0) or 0),
                        "connection_size": str(row.get("connection_size", "")) if pd.notna(row.get("connection_size")) else "",
                        "remarks": str(row.get("remarks", "")) if pd.notna(row.get("remarks")) else "",
                        "remarks_marathi": str(row.get("remarks_marathi", "")) if pd.notna(row.get("remarks_marathi")) else "",
                    }


                    # Create survey
                    survey_serializer = SurveySerializer(data=survey_data, context={"request": request})
                    if survey_serializer.is_valid():
                        survey_serializer.save()
                        success_count += 1
                        logger.info(f"Successfully imported Ward {ward_no}, Property {property_no}")
                    else:
                        error_count += 1
                        errors.append(
                            f"Row {index+2}: Ward {ward_no}, Property {property_no} - {survey_serializer.errors}"
                        )

                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {index+2}: {str(e)}")
                    logger.error(f"Error importing row {index+2}: {str(e)}")

        return {
            "success": True,
            "message": f"Import completed: {success_count} successful, {error_count} errors",
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors[:10] if errors else [],
        }


# ==============================
# Excel Template Download
# ==============================
@method_decorator(csrf_exempt, name="dispatch")
class SurveyExcelTemplateDownloadView(APIView):
    """Download Excel import template file"""

    permission_classes = [IsAuthenticated, HasModuleAccess]
    required_permission = "download-survey-template"

    def get(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Survey_Import_Template"

            # Headers
            headers = [
                "ward_no",
                "property_no",
                "old_ward_no",
                "old_property_no",
                "property_description",
                "address",
                "address_marathi",
                "water_connection_available",
                "number_of_water_connections",
                "connection_size",
                "remarks",
                "remarks_marathi",
            ]

            ws.append(headers)

            # Sample data
            sample_data = [
                [
                    1,  # ward_no
                    101,  # property_no
                    "",  # old_ward_no
                    "",  # old_property_no
                    "Residential Building",  # property_description
                    "Plot No 101, Sector A",  # address
                    "प्लॉट नं १०१, सेक्टर ए",  # address_marathi
                    "Yes",  # water_connection_available
                    1,  # number_of_water_connections
                    "15mm",  # connection_size
                    "Well maintained property",  # remarks
                    "चांगली मालमत्ता",  # remarks_marathi
                ],
            ]

            for row_data in sample_data:
                ws.append(row_data)

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create instructions sheet
            ws_instructions = wb.create_sheet("Instructions")

            instructions = [
                ["Survey Import Template Instructions", "", ""],
                ["", "", ""],
                ["Step", "English", "Marathi"],
                ["1", "Fill the required fields marked with *", "* ने चिन्हांकित आवश्यक फील्ड भरा"],
                ["2", "ward_no* - Ward number (integer)", "वार्ड क्रमांक (पूर्णांक)"],
                ["3", "property_no* - Property number", "मालमत्ता क्रमांक"],
                ["4", "water_connection_available: Yes/No", "पाणी कनेक्शन: होय/नाही"],
                ["5", "Save as .xlsx and upload via Import Excel", "xlsx म्हणून सेव्ह करा आणि Import Excel द्वारे अपलोड करा"],
            ]

            for row in instructions:
                ws_instructions.append(row)

            # Style instructions header
            for cell in ws_instructions[1]:
                cell.font = Font(bold=True, size=14)
                cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")

            # Save to BytesIO
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Create response
            filename = f"Survey_Import_Template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            logger.info(f"Excel template downloaded: {filename}")
            return response

        except Exception as e:
            logger.error(f"Template download error: {str(e)}")
            return Response(
                {"success": False, "message": f"Template download error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ==============================
# Survey Statistics API
# ==============================
class SurveyStatsView(APIView):
    """Get Survey Statistics for Dashboard"""
    
    permission_classes = [IsAuthenticated, HasModuleAccess]
    required_permission = "access-survey"
    
    def get(self, request):
        try:
            # 1. TOTAL COUNTS
            total_surveys = Survey.objects.count()
            total_wards = Survey.objects.values('ward_no').distinct().count()
            
            # 2. THIS MONTH'S SURVEYS
            current_month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            this_month_surveys = Survey.objects.filter(
                created_at__gte=current_month_start
            ).count()
            
            # 3. LAST MONTH'S SURVEYS (for percentage change)
            last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
            last_month_surveys = Survey.objects.filter(
                created_at__gte=last_month_start,
                created_at__lt=current_month_start
            ).count()
            
            # Calculate percentage change
            if last_month_surveys > 0:
                month_change = ((this_month_surveys - last_month_surveys) / last_month_surveys) * 100
            else:
                month_change = 100 if this_month_surveys > 0 else 0
            
            # 4. MONTHLY SURVEY TREND (Last 12 months)
            twelve_months_ago = datetime.now() - timedelta(days=365)
            monthly_surveys = Survey.objects.filter(
                created_at__gte=twelve_months_ago
            ).annotate(
                month=TruncMonth('created_at')
            ).values('month').annotate(
                count=Count('id')
            ).order_by('month')
            
            # 5. WARD-WISE SURVEY COUNT (Top 10)
            ward_wise_count = Survey.objects.values('ward_no').annotate(
                count=Count('id')
            ).order_by('-count')[:10]
            
            # 6. PROPERTY DESCRIPTION DISTRIBUTION
            property_types = Survey.objects.values('property_description').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # 7. WARD PROGRESS (for each ward with target)
            # Assuming each ward has a target of 100 surveys
            TARGET_PER_WARD = 100
            ward_progress = []
            
            all_wards = Survey.objects.values('ward_no').distinct().order_by('ward_no')
            for ward in all_wards:
                ward_no = ward['ward_no']
                survey_count = Survey.objects.filter(ward_no=ward_no).count()
                progress_percentage = min((survey_count / TARGET_PER_WARD) * 100, 100)
                
                ward_progress.append({
                    'ward_no': ward_no,
                    'completed': survey_count,
                    'target': TARGET_PER_WARD,
                    'progress_percentage': round(progress_percentage, 2)
                })
            
            return Response({
                'success': True,
                'data': {
                    # Metrics
                    'total_surveys': total_surveys,
                    'total_wards': total_wards,
                    'this_month_surveys': this_month_surveys,
                    'last_month_surveys': last_month_surveys,
                    'month_change_percentage': round(month_change, 2),
                    
                    # Monthly Trend Chart
                    'monthly_trend': [
                        {
                            'month': item['month'].strftime('%b %Y'),
                            'count': item['count']
                        }
                        for item in monthly_surveys
                    ],
                    
                    # Ward-wise Distribution (Bar Chart)
                    'ward_wise_distribution': [
                        {
                            'ward_no': item['ward_no'],
                            'count': item['count']
                        }
                        for item in ward_wise_count
                    ],
                    
                    # Property Type Distribution (Pie Chart)
                    'property_type_distribution': [
                        {
                            'type': item['property_description'] or 'Not Specified',
                            'count': item['count']
                        }
                        for item in property_types
                    ],
                    
                    # Ward Progress
                    'ward_progress': ward_progress[:5]  # Top 5 wards
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching survey statistics: {str(e)}")
            return Response(
                {
                    'success': False,
                    'message': f'Error fetching statistics: {str(e)}'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
